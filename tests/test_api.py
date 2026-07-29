import os
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

ROOT = Path(__file__).resolve().parents[1]
os.environ["DATABASE_URL"] = "sqlite+pysqlite:////tmp/pnc-test.db"
os.environ["ADMIN_API_KEY"] = "test-key"
os.environ["SEED_CSV_PATH"] = str(ROOT / "data" / "seed" / "institutions.csv")
os.environ["STORAGE_DIR"] = str(ROOT / "storage")
os.environ["CRAWL_ALLOWLIST"] = "example.com,localhost"

from app.config import get_settings
from app.db import Base, engine, init_db
from app.main import create_app
from app.services.collectors.html_generic import fetch_and_parse, load_fixture
from app.services.normalize import normalize_name, normalize_phone


@pytest.fixture(autouse=True)
def _fresh_db():
    get_settings.cache_clear()
    Base.metadata.drop_all(bind=engine)
    init_db()
    yield
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def client():
    app = create_app()
    with TestClient(app) as c:
        yield c


def test_normalize():
    assert normalize_phone("8 (495) 123-45-67") == "74951234567"
    assert "перинатальный" in normalize_name('ГБУЗ «Перинатальный центр»')


def test_health_and_list(client):
    r = client.get("/api/v1/health")
    assert r.status_code == 200
    r = client.get("/api/v1/institutions?page_size=5")
    assert r.status_code == 200
    data = r.json()
    assert data["total"] >= 100
    assert len(data["items"]) == 5


def test_filter_search(client):
    r = client.get("/api/v1/institutions", params={"region": "Москва", "page_size": 50})
    assert r.status_code == 200
    assert r.json()["total"] >= 1
    r = client.get("/api/v1/institutions", params={"q": "НМИЦ", "page_size": 10})
    assert r.json()["total"] >= 1
    r = client.get("/api/v1/institutions", params={"has_email": True, "page_size": 10})
    assert all(i["emails"] for i in r.json()["items"])


def test_admin_auth_and_export(client):
    r = client.post("/api/v1/admin/export", json={})
    assert r.status_code == 401
    r = client.post("/api/v1/admin/export", json={"region": "Москва"}, headers={"X-API-Key": "test-key"})
    assert r.status_code == 202
    job = r.json()
    assert job["status"] == "done"
    file_r = client.get(f"/api/v1/admin/export/{job['id']}/file", headers={"X-API-Key": "test-key"})
    assert file_r.status_code == 200
    assert file_r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_mailing_dry_run(client):
    r = client.post(
        "/api/v1/admin/mailings",
        json={"subject": "t", "body_html": "<p>x</p>", "dry_run": True, "filter": {"has_email": True}},
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 202
    body = r.json()
    assert body["status"] == "done"
    assert body["skipped_count"] > 0
    assert body["sent_count"] == 0


def test_html_collector_fixture():
    html = load_fixture(ROOT / "data" / "fixtures" / "sample_contacts.html")
    parsed = fetch_and_parse("https://example.com/contacts", html_override=html)
    assert "info@example.com" in parsed["emails"]
    assert any(p.endswith("4951234567") or "74951234567" in p for p in parsed["phones"])


def test_ssrf_blocked():
    with pytest.raises(ValueError):
        fetch_and_parse("https://evil.example.net/x")


def test_crawl_2gis_requires_key(client):
    r = client.post(
        "/api/v1/admin/jobs/crawl",
        json={"source": "2gis", "cities": ["Москва"]},
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 202
    body = r.json()
    assert body["status"] == "failed"
    assert "DGIS_API_KEY" in (body.get("error") or "")


def test_persons_seed_is_loaded_on_startup(client):
    """Снапшот персон должен восстанавливаться, иначе свежий деплой теряет должности."""
    r = client.get("/api/v1/admin/persons", params={"role": "chief", "limit": 5}, headers={"X-API-Key": "test-key"})
    assert r.status_code == 200
    assert r.json()["total"] >= 1


def test_persons_crud_and_field_sync(client):
    inst = client.get("/api/v1/institutions?page_size=1").json()["items"][0]
    person = _add_person(inst["id"], "Иванов Иван Иванович", "chief", "high")

    r = client.get(f"/api/v1/institutions/{inst['id']}/persons")
    assert r.status_code == 200
    assert any(p["full_name"] == "Иванов Иван Иванович" for p in r.json()["items"])

    r = client.patch(
        f"/api/v1/admin/persons/{person}",
        json={"full_name": "Иванова Мария Петровна", "verified_manually": True},
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 200
    assert r.json()["confidence"] == "high"

    detail = client.get(f"/api/v1/institutions/{inst['id']}").json()
    assert detail["chief_physician"] == "Иванова Мария Петровна"

    r = client.delete(f"/api/v1/admin/persons/{person}", headers={"X-API-Key": "test-key"})
    assert r.status_code == 204
    detail = client.get(f"/api/v1/institutions/{inst['id']}").json()
    assert detail["chief_physician"] is None


def test_persons_low_confidence_hidden_by_default(client):
    inst = client.get("/api/v1/institutions?page_size=1").json()["items"][0]
    _add_person(inst["id"], "Слабый Кандидат Иванович", "chief", "low")
    items = client.get(f"/api/v1/institutions/{inst['id']}/persons").json()["items"]
    assert not any(p["confidence"] == "low" for p in items)
    items = client.get(
        f"/api/v1/institutions/{inst['id']}/persons", params={"min_confidence": "low"}
    ).json()["items"]
    assert any(p["confidence"] == "low" for p in items)


def test_persons_endpoint_requires_admin_key(client):
    assert client.get("/api/v1/admin/persons").status_code == 401
    assert client.get("/api/v1/admin/persons", headers={"X-API-Key": "test-key"}).status_code == 200


def test_completeness_metrics(client):
    r = client.get("/api/v1/admin/metrics/completeness", headers={"X-API-Key": "test-key"})
    assert r.status_code == 200
    data = r.json()
    assert data["total"] > 0
    assert "chief_physician" in data["fields"]
    assert data["fields"]["address"]["pct"] > 50
    assert isinstance(data["by_type"], list) and data["by_type"]


def test_has_chief_filter(client):
    inst = client.get("/api/v1/institutions?page_size=1").json()["items"][0]
    client.patch(
        f"/api/v1/admin/institutions/{inst['id']}",
        json={"chief_physician": "Тестов Тест Тестович"},
        headers={"X-API-Key": "test-key"},
    )
    with_chief = client.get("/api/v1/institutions", params={"has_chief": True, "page_size": 100}).json()
    assert with_chief["total"] >= 1
    assert all(i["chief_physician"] for i in with_chief["items"])
    without = client.get("/api/v1/institutions", params={"has_chief": False, "page_size": 100}).json()
    assert all(not i["chief_physician"] for i in without["items"])


def test_mailing_preview_personalizes(client):
    inst = client.get("/api/v1/institutions", params={"has_email": True, "page_size": 1}).json()["items"][0]
    client.patch(
        f"/api/v1/admin/institutions/{inst['id']}",
        json={"chief_physician": "Тестов Тест Тестович"},
        headers={"X-API-Key": "test-key"},
    )
    r = client.post(
        "/api/v1/admin/mailings/preview",
        json={
            "subject": "Приглашение для {{name}}",
            "body_html": "<p>Уважаемый(ая) {{chief}}</p>",
            "filter": {"has_email": True},
        },
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["total_recipients"] > 0
    assert data["samples"]
    assert all("{{" not in s["body_html"] for s in data["samples"])


def test_enrich_job_is_recorded(client):
    r = client.post(
        "/api/v1/admin/jobs/enrich",
        json={"limit": 1, "region": "НетТакогоРегиона"},
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 202
    job = r.json()
    assert job["kind"] == "enrich"
    assert job["status"] == "done"
    assert job["result_json"]["processed"] == 0
    assert client.get(f"/api/v1/admin/jobs/{job['id']}", headers={"X-API-Key": "test-key"}).status_code == 200


def test_enrich_by_institution_id_targets_exactly_that_row(client, monkeypatch):
    """Кнопка «обогатить это учреждение» не должна обходить чужой сайт."""
    from app.services import enrich as enrich_module

    items = client.get("/api/v1/institutions", params={"region": "Москва", "page_size": 5}).json()["items"]
    target = items[-1]
    processed: list[str] = []

    def fake_enrich(db, inst, *, force=False):
        processed.append(inst.id)
        return {"id": inst.id, "name": inst.name, "status": "ok", "persons": 0}

    monkeypatch.setattr(enrich_module, "enrich_institution", fake_enrich)
    r = client.post(
        "/api/v1/admin/jobs/enrich",
        # регион и лимит намеренно противоречат адресному запуску
        json={"institution_id": target["id"], "region": "Другой регион", "limit": 25, "force": True},
        headers={"X-API-Key": "test-key"},
    )
    assert r.status_code == 202
    assert r.json()["result_json"]["processed"] == 1
    assert processed == [target["id"]]


def test_enrich_keeps_better_stored_chief(client):
    """Повторный обход, нашедший только medium-кандидата, не затирает high из БД."""
    from app.db import SessionLocal
    from app.models import Institution
    from app.services.enrich import sync_institution_fields

    inst = client.get("/api/v1/institutions?page_size=1").json()["items"][0]
    _add_person(inst["id"], "Достоверный Кандидат Иванович", "chief", "high")
    _add_person(inst["id"], "Сомнительный Кандидат Петрович", "chief", "medium")

    with SessionLocal() as db:
        row = db.get(Institution, inst["id"])
        sync_institution_fields(db, row)
        db.commit()
        assert row.chief_physician == "Достоверный Кандидат Иванович"


def test_export_has_persons_sheet(client):
    from openpyxl import load_workbook

    inst = client.get("/api/v1/institutions", params={"region": "Москва", "page_size": 1}).json()["items"][0]
    _add_person(inst["id"], "Экспортов Экспорт Экспортович", "chief", "high")
    r = client.post("/api/v1/admin/export", json={"region": "Москва"}, headers={"X-API-Key": "test-key"})
    path = r.json()["result_json"]["path"]
    wb = load_workbook(path)
    assert "persons" in wb.sheetnames
    values = [c.value for row in wb["persons"].iter_rows() for c in row]
    assert "Экспортов Экспорт Экспортович" in values


def _add_person(institution_id: str, full_name: str, role: str, confidence: str) -> str:
    from app.db import SessionLocal
    from app.models import InstitutionPerson
    from app.services.collectors.person_extractor import normalize_person_name

    with SessionLocal() as db:
        person = InstitutionPerson(
            institution_id=institution_id,
            full_name=full_name,
            full_name_norm=normalize_person_name(full_name),
            role=role,
            confidence=confidence,
            source_url="https://example.com/rukovodstvo",
        )
        db.add(person)
        db.commit()
        return person.id


def test_classify_and_sites_registry():
    from app.services.collectors.base import classify_type
    from app.services.collectors.site_registry import load_registry

    assert classify_type("Женская консультация №5") == "womens_clinic"
    assert classify_type("Областной перинатальный центр") == "perinatal_center_regional"
    rows = load_registry(ROOT / "data" / "registry" / "official_sites.yaml")
    assert len(rows) >= 4
    assert any("Кулакова" in r["name"] for r in rows)
